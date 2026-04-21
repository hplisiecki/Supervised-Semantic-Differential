"""Tests for ssdiff.results.paired_view — Task 3.

Exercises all save formats and the paired-view interface using
hand-constructed views (no real SSD fit needed).
"""

from __future__ import annotations

import json
import warnings
from pathlib import Path

import pytest

from ssdiff.results.continuous_result import ClustersViewSided, SnippetsView, WordsView
from ssdiff.results.paired_view import (
    ClustersViewPaired,
    ClustersViewSidedPaired,
    SnippetsViewPaired,
    WordsViewPaired,
)
from ssdiff.results.schema import Cluster, Snippet, Word


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


def _make_words_view(contrast: str) -> WordsView:
    return WordsView([
        Word(side="pos", rank=1, word=f"{contrast}_p", cos_beta=0.9, contrast=contrast),
        Word(side="neg", rank=1, word=f"{contrast}_n", cos_beta=-0.9, contrast=contrast),
    ])


def _make_snippets_view(contrast: str) -> SnippetsView:
    return SnippetsView([
        Snippet(
            snippet_id=0, side="pos", doc_id=0, cosine=0.8,
            seed="test", start_token_idx=0, end_token_idx=3,
            start_sent_idx=0, end_sent_idx=1,
            text_window="hello world", text_surface="hello world end",
            text_lemmas="hello world end", contrast=contrast,
        ),
    ])


def _make_clusters_view_sided(contrast: str, side: str) -> ClustersViewSided:
    return ClustersViewSided(
        parent=None,
        side=side,
        rows=[
            Cluster(
                cluster_id=0, side=side, size=3,
                coherence=0.7, centroid_cos_beta=0.8,
                contrast=contrast,
            )
        ],
        words_rows=[],
        params={},
    )


class _FakePairClustersIndex:
    """Minimal stand-in for _PairClustersIndex to back ClustersViewPaired."""

    def __init__(self, contrast: str):
        self._contrast = contrast

    @property
    def pos(self) -> ClustersViewSided:
        return _make_clusters_view_sided(self._contrast, "pos")

    @property
    def neg(self) -> ClustersViewSided:
        return _make_clusters_view_sided(self._contrast, "neg")


@pytest.fixture
def paired_words() -> WordsViewPaired:
    return WordsViewPaired(
        views={
            ("g1", "g2"): _make_words_view("g1_g2"),
            ("g1", "g3"): _make_words_view("g1_g3"),
            ("g2", "g3"): _make_words_view("g2_g3"),
        },
        view_name="words",
    )


@pytest.fixture
def paired_snippets() -> SnippetsViewPaired:
    return SnippetsViewPaired(
        views={
            ("g1", "g2"): _make_snippets_view("g1_g2"),
            ("g1", "g3"): _make_snippets_view("g1_g3"),
            ("g2", "g3"): _make_snippets_view("g2_g3"),
        },
        view_name="snippets",
    )


@pytest.fixture
def paired_clusters_index() -> ClustersViewPaired:
    return ClustersViewPaired(
        views={
            ("g1", "g2"): _FakePairClustersIndex("g1_g2"),
            ("g1", "g3"): _FakePairClustersIndex("g1_g3"),
            ("g2", "g3"): _FakePairClustersIndex("g2_g3"),
        },
    )


# ---------------------------------------------------------------------------
# Interface tests
# ---------------------------------------------------------------------------


def test_keys_and_len(paired_words: WordsViewPaired) -> None:
    assert sorted(paired_words.keys()) == [("g1", "g2"), ("g1", "g3"), ("g2", "g3")]
    # len() counts total rows across all pairs (2 rows x 3 pairs = 6)
    assert len(paired_words) == 6


def test_getitem_canonical(paired_words: WordsViewPaired) -> None:
    view = paired_words[("g1", "g2")]
    assert isinstance(view, WordsView)
    assert len(view) == 2


def test_getitem_reverse_order_raises(paired_words: WordsViewPaired) -> None:
    with pytest.raises(KeyError):
        _ = paired_words[("g2", "g1")]


def test_getitem_unknown_key_raises(paired_words: WordsViewPaired) -> None:
    with pytest.raises(KeyError):
        _ = paired_words[("g1", "g9")]


def test_getitem_bad_type_raises(paired_words: WordsViewPaired) -> None:
    with pytest.raises(KeyError):
        _ = paired_words["g1_g2"]  # type: ignore[index]


def test_iter_yields_rows(paired_words: WordsViewPaired) -> None:
    """Flat iteration yields Word rows (not (key, view) tuples) — new Phase C contract."""
    from ssdiff.results.schema import Word
    items = list(paired_words)
    assert len(items) == 6  # 2 words per pair × 3 pairs
    assert all(isinstance(item, Word) for item in items)
    contrasts = {item.contrast for item in items}
    assert contrasts == {"g1_g2", "g1_g3", "g2_g3"}


def test_repr_contains_view_name(paired_words: WordsViewPaired) -> None:
    r = repr(paired_words)
    assert "words" in r
    assert "n=3" in r


# ---------------------------------------------------------------------------
# Save — csv fan-out (N≥2)
# ---------------------------------------------------------------------------


def test_save_csv_fans_out(paired_words: WordsViewPaired, tmp_path: Path) -> None:
    target = tmp_path / "anything.csv"
    with warnings.catch_warnings(record=True) as recorded:
        warnings.simplefilter("always")
        paired_words.save(target)
    # Original path NOT written
    assert not target.exists()
    # Subfolder hardcoded to view_name "words", not "anything"
    folder = tmp_path / "words"
    assert folder.is_dir()
    assert (folder / "g1_g2.csv").is_file()
    assert (folder / "g1_g3.csv").is_file()
    assert (folder / "g2_g3.csv").is_file()
    assert any(issubclass(w.category, UserWarning) for w in recorded)


def test_save_csv_content_correct(paired_words: WordsViewPaired, tmp_path: Path) -> None:
    with warnings.catch_warnings(record=True):
        warnings.simplefilter("always")
        paired_words.save(tmp_path / "out.csv")
    content = (tmp_path / "words" / "g1_g2.csv").read_text()
    assert "g1_g2_p" in content
    assert "side" in content


# ---------------------------------------------------------------------------
# Save — xlsx multi-sheet (N≥2)
# ---------------------------------------------------------------------------


def test_save_xlsx_multisheet(paired_words: WordsViewPaired, tmp_path: Path) -> None:
    path = tmp_path / "out.xlsx"
    paired_words.save(path)
    assert path.is_file()
    from openpyxl import load_workbook
    wb = load_workbook(path)
    assert set(wb.sheetnames) == {"g1_g2", "g1_g3", "g2_g3"}


def test_save_xlsx_row_content(paired_words: WordsViewPaired, tmp_path: Path) -> None:
    path = tmp_path / "out.xlsx"
    paired_words.save(path)
    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb["g1_g2"]
    # Row 1 = header, Row 2+ = data
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    assert "word" in headers


# ---------------------------------------------------------------------------
# Save — json keyed (N≥2)
# ---------------------------------------------------------------------------


def test_save_json_keyed(paired_words: WordsViewPaired, tmp_path: Path) -> None:
    path = tmp_path / "out.json"
    paired_words.save(path)
    data = json.loads(path.read_text())
    assert set(data.keys()) == {"g1_g2", "g1_g3", "g2_g3"}
    for rows in data.values():
        assert isinstance(rows, list)
        assert all(isinstance(r, dict) for r in rows)


def test_save_json_contains_words(paired_words: WordsViewPaired, tmp_path: Path) -> None:
    path = tmp_path / "out.json"
    paired_words.save(path)
    data = json.loads(path.read_text())
    pair_rows = data["g1_g2"]
    assert any(r.get("word") == "g1_g2_p" for r in pair_rows)


# ---------------------------------------------------------------------------
# Save — md / html / tex / docx / txt sectioned (N≥2)
# ---------------------------------------------------------------------------


def test_save_md_sectioned(paired_words: WordsViewPaired, tmp_path: Path) -> None:
    path = tmp_path / "out.md"
    paired_words.save(path)
    body = path.read_text()
    assert "g1 vs g2" in body
    assert "g1 vs g3" in body
    assert "g2 vs g3" in body


def test_save_html_sectioned(paired_words: WordsViewPaired, tmp_path: Path) -> None:
    path = tmp_path / "out.html"
    paired_words.save(path)
    body = path.read_text()
    assert "g1 vs g2" in body
    assert "<table>" in body


def test_save_tex_sectioned(paired_words: WordsViewPaired, tmp_path: Path) -> None:
    path = tmp_path / "out.tex"
    paired_words.save(path)
    body = path.read_text()
    assert "g1 vs g2" in body
    assert "tabular" in body


def test_save_txt_sectioned(paired_words: WordsViewPaired, tmp_path: Path) -> None:
    path = tmp_path / "out.txt"
    paired_words.save(path)
    body = path.read_text()
    assert "g1 vs g2" in body
    assert "g2 vs g3" in body


def test_save_docx_sectioned(paired_words: WordsViewPaired, tmp_path: Path) -> None:
    path = tmp_path / "out.docx"
    paired_words.save(path)
    assert path.is_file()
    # Verify it's a valid docx with correct headings
    from docx import Document
    doc = Document(str(path))
    headings = [p.text for p in doc.paragraphs if p.style.name.startswith("Heading")]
    assert "g1 vs g2" in headings
    assert "g2 vs g3" in headings


# ---------------------------------------------------------------------------
# Save — N=1 single-pair no fan-out
# ---------------------------------------------------------------------------


def test_save_single_pair_no_fanout(tmp_path: Path) -> None:
    single = WordsViewPaired(
        views={
            ("g1", "g2"): _make_words_view("g1_g2"),
        },
        view_name="words",
    )
    target = tmp_path / "w.csv"
    single.save(target)
    # N=1: flat file at target, no subfolder
    assert target.is_file()
    assert not (tmp_path / "words").exists()


def test_save_single_pair_xlsx_flat(tmp_path: Path) -> None:
    single = WordsViewPaired(
        views={("g1", "g2"): _make_words_view("g1_g2")},
        view_name="words",
    )
    path = tmp_path / "w.xlsx"
    single.save(path)
    assert path.is_file()


# ---------------------------------------------------------------------------
# Save — default path (no path arg)
# ---------------------------------------------------------------------------


def test_save_default_path_creates_csv(
    paired_words: WordsViewPaired, tmp_path: Path, monkeypatch
) -> None:
    """Default save with no path → <cwd>/<view_name>.csv (fan-out for N≥2)."""
    monkeypatch.chdir(tmp_path)
    with warnings.catch_warnings(record=True):
        warnings.simplefilter("always")
        paired_words.save()
    # N≥2 + csv → fan-out; subfolder is "words"
    folder = tmp_path / "words"
    assert folder.is_dir()
    assert (folder / "g1_g2.csv").is_file()


def test_save_default_path_single_pair(tmp_path: Path, monkeypatch) -> None:
    single = WordsViewPaired(
        views={("g1", "g2"): _make_words_view("g1_g2")},
        view_name="words",
    )
    monkeypatch.chdir(tmp_path)
    single.save()
    assert (tmp_path / "words.csv").is_file()


# ---------------------------------------------------------------------------
# ClustersViewPaired — pos / neg access + save
# ---------------------------------------------------------------------------


def test_clusters_index_paired_keys(
    paired_clusters_index: ClustersViewPaired,
) -> None:
    assert sorted(paired_clusters_index.keys()) == [
        ("g1", "g2"),
        ("g1", "g3"),
        ("g2", "g3"),
    ]
    # len() counts total Cluster rows across all pairs (1 pos + 1 neg per pair × 3 pairs = 6)
    assert len(paired_clusters_index) == 6


def test_clusters_index_paired_pos_type(
    paired_clusters_index: ClustersViewPaired,
) -> None:
    pos = paired_clusters_index.pos
    assert isinstance(pos, ClustersViewSidedPaired)
    assert pos._view_name == "clusters_pos"


def test_clusters_index_paired_neg_type(
    paired_clusters_index: ClustersViewPaired,
) -> None:
    neg = paired_clusters_index.neg
    assert isinstance(neg, ClustersViewSidedPaired)
    assert neg._view_name == "clusters_neg"


def test_clusters_pos_save_subfolder_name(
    paired_clusters_index: ClustersViewPaired, tmp_path: Path
) -> None:
    """clusters.pos.save → subfolder name = 'clusters_pos', NOT 'anything'."""
    target = tmp_path / "anything.csv"
    with warnings.catch_warnings(record=True) as recorded:
        warnings.simplefilter("always")
        paired_clusters_index.pos.save(target)
    assert not target.exists()
    folder = tmp_path / "clusters_pos"
    assert folder.is_dir()
    assert (folder / "g1_g2.csv").is_file()
    assert any(issubclass(w.category, UserWarning) for w in recorded)


def test_clusters_neg_save_subfolder_name(
    paired_clusters_index: ClustersViewPaired, tmp_path: Path
) -> None:
    target = tmp_path / "anything.csv"
    with warnings.catch_warnings(record=True):
        warnings.simplefilter("always")
        paired_clusters_index.neg.save(target)
    folder = tmp_path / "clusters_neg"
    assert folder.is_dir()
    assert (folder / "g1_g2.csv").is_file()


def test_clusters_pos_save_xlsx(
    paired_clusters_index: ClustersViewPaired, tmp_path: Path
) -> None:
    path = tmp_path / "clusters.xlsx"
    paired_clusters_index.pos.save(path)
    assert path.is_file()
    from openpyxl import load_workbook
    wb = load_workbook(path)
    assert set(wb.sheetnames) == {"g1_g2", "g1_g3", "g2_g3"}


def test_clusters_index_paired_reverse_raises(
    paired_clusters_index: ClustersViewPaired,
) -> None:
    with pytest.raises(KeyError):
        _ = paired_clusters_index[("g2", "g1")]


def test_clusters_index_paired_iter(
    paired_clusters_index: ClustersViewPaired,
) -> None:
    """Flat iteration yields Cluster rows (not (key, view) tuples) — new Phase C contract."""
    from ssdiff.results.schema import Cluster
    items = list(paired_clusters_index)
    assert len(items) == 6  # 1 pos + 1 neg per pair × 3 pairs
    assert all(isinstance(item, Cluster) for item in items)


# ---------------------------------------------------------------------------
# SnippetsViewPaired
# ---------------------------------------------------------------------------


def test_snippets_view_paired_keys(
    paired_snippets: SnippetsViewPaired,
) -> None:
    assert sorted(paired_snippets.keys()) == [
        ("g1", "g2"),
        ("g1", "g3"),
        ("g2", "g3"),
    ]


def test_snippets_save_csv_fanout(
    paired_snippets: SnippetsViewPaired, tmp_path: Path
) -> None:
    target = tmp_path / "anything.csv"
    with warnings.catch_warnings(record=True) as recorded:
        warnings.simplefilter("always")
        paired_snippets.save(target)
    folder = tmp_path / "snippets"
    assert folder.is_dir()
    assert (folder / "g1_g2.csv").is_file()
    assert any(issubclass(w.category, UserWarning) for w in recorded)


def test_snippets_save_json(
    paired_snippets: SnippetsViewPaired, tmp_path: Path
) -> None:
    path = tmp_path / "out.json"
    paired_snippets.save(path)
    data = json.loads(path.read_text())
    assert set(data.keys()) == {"g1_g2", "g1_g3", "g2_g3"}


def test_snippets_save_md(
    paired_snippets: SnippetsViewPaired, tmp_path: Path
) -> None:
    path = tmp_path / "out.md"
    paired_snippets.save(path)
    body = path.read_text()
    assert "g1 vs g2" in body


# ---------------------------------------------------------------------------
# cols / k filtering
# ---------------------------------------------------------------------------


def test_save_with_k_filter(paired_words: WordsViewPaired, tmp_path: Path) -> None:
    """k=1 caps each view to 1 row."""
    with warnings.catch_warnings(record=True):
        warnings.simplefilter("always")
        paired_words.save(tmp_path / "out.csv", k=1)
    content = (tmp_path / "words" / "g1_g2.csv").read_text()
    lines = [l for l in content.strip().splitlines() if l]
    # header + 1 data row
    assert len(lines) == 2


def test_save_json_with_cols(paired_words: WordsViewPaired, tmp_path: Path) -> None:
    path = tmp_path / "out.json"
    paired_words.save(path, cols=["word", "cos_beta"])
    data = json.loads(path.read_text())
    rows = data["g1_g2"]
    assert all(set(r.keys()) == {"word", "cos_beta"} for r in rows)


# ---------------------------------------------------------------------------
# Unsupported extension raises
# ---------------------------------------------------------------------------


def test_save_unsupported_ext_raises(
    paired_words: WordsViewPaired, tmp_path: Path
) -> None:
    with pytest.raises(ValueError):
        paired_words.save(tmp_path / "out.xyz")


# ---------------------------------------------------------------------------
# Warning behaviour: fan-out fires exactly once and points to user code
# ---------------------------------------------------------------------------


def test_save_csv_warning_fires_once(paired_words: WordsViewPaired, tmp_path: Path) -> None:
    with warnings.catch_warnings(record=True) as rec:
        warnings.simplefilter("always")
        paired_words.save(tmp_path / "out.csv")
    fan_out_warnings = [w for w in rec if "fans out" in str(w.message) or "multi-pair csv" in str(w.message).lower()]
    assert len(fan_out_warnings) == 1, f"expected exactly 1 fan-out warning, got {len(fan_out_warnings)}: {[str(w.message) for w in fan_out_warnings]}"


def test_save_csv_warning_stacklevel(paired_words: WordsViewPaired, tmp_path: Path) -> None:
    with warnings.catch_warnings(record=True) as rec:
        warnings.simplefilter("always")
        paired_words.save(tmp_path / "out.csv")  # <-- this file is the caller
    fan_out = [w for w in rec if issubclass(w.category, UserWarning)]
    assert fan_out
    # Warning filename should be this test file, not paired_view.py
    assert "paired_view.py" not in fan_out[0].filename, (
        f"stacklevel wrong: warning points to library internals {fan_out[0].filename}"
    )


# ---------------------------------------------------------------------------
# Import smoke test
# ---------------------------------------------------------------------------


def test_imports() -> None:
    from ssdiff.results.paired_view import (  # noqa: F401
        ClustersViewPaired,
        ClustersViewSidedPaired,
        SnippetsViewPaired,
        WordsViewPaired,
    )
